from __future__ import annotations

from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "docs" / "YAM_AGRI_WBS_GANTT.xlsx",
    ROOT / "docs" / "planning" / "YAM_AGRI_WBS_GANTT.xlsx",
]

STATUS_BY_MILESTONE = {
    "M0": "✅ Done",
    "M1": "✅ Done",
    "M2": "🟨 In Progress",
    "M3": "⬜ Pending",
    "M4": "⬜ Pending",
    "M5": "⬜ Pending",
    "M6": "⬜ Pending",
    "M7": "⬜ Pending",
    "M8": "⬜ Pending",
    "M9": "⬜ Pending",
}

for path in FILES:
    wb = load_workbook(path)
    ws = wb["Milestones"]

    headers = {str(cell.value).strip(): idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
    col_milestone = headers.get("Milestone")
    col_status = headers.get("Status")
    if not col_milestone or not col_status:
        raise RuntimeError(f"Milestones sheet missing required columns in {path}")

    for r in range(2, ws.max_row + 1):
        milestone = ws.cell(row=r, column=col_milestone).value
        if milestone in STATUS_BY_MILESTONE:
            ws.cell(row=r, column=col_status).value = STATUS_BY_MILESTONE[milestone]

    wb.save(path)
    print(f"updated: {path}")
