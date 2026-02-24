from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "docs" / "YAM_AGRI_WBS_GANTT.xlsx",
    ROOT / "docs" / "planning" / "YAM_AGRI_WBS_GANTT.xlsx",
]

STATUS_DONE = "Done"
STATUS_PARTIAL = "Partial"
STATUS_NOT_STARTED = "Not Started"

CATEGORY_COLUMNS = [
    "Schema",
    "Validation",
    "Permissions/Isolation",
    "Workflow",
    "Evidence",
]

MILESTONE_TO_PHASE = {
    "M0": "0",
    "M1": "1",
    "M2": "2",
    "M3": "3",
    "M4": "4",
    "M5": "5",
    "M6": "6",
    "M7": "7",
    "M8": "8",
    "M9": "9",
}


def _header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(ws[1])
        if cell.value is not None and str(cell.value).strip()
    }


def _ensure_column(ws, headers: dict[str, int], header_name: str) -> int:
    if header_name in headers:
        return headers[header_name]

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header_name
    headers[header_name] = col
    return col


def _normalize_milestone_status(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return STATUS_NOT_STARTED
    if "âœ…" in text or "done" in text:
        return STATUS_DONE
    if "ðŸŸ¨" in text or "in progress" in text or "progress" in text:
        return STATUS_PARTIAL
    return STATUS_NOT_STARTED


def _phase_status_from_milestones(ws_milestones) -> dict[str, dict[str, object]]:
    headers = _header_map(ws_milestones)
    col_milestone = headers.get("Milestone")
    col_status = headers.get("Status")
    col_target_date = headers.get("Target Date")
    if not col_milestone or not col_status:
        raise RuntimeError("Milestones sheet missing Milestone/Status columns")

    phase_map: dict[str, dict[str, object]] = {}
    for r in range(2, ws_milestones.max_row + 1):
        milestone = ws_milestones.cell(row=r, column=col_milestone).value
        if not milestone:
            continue

        milestone_key = str(milestone).strip().upper()
        phase = MILESTONE_TO_PHASE.get(milestone_key)
        if not phase:
            continue

        normalized_status = _normalize_milestone_status(ws_milestones.cell(row=r, column=col_status).value)
        target_date = ws_milestones.cell(row=r, column=col_target_date).value if col_target_date else None
        phase_map[phase] = {
            "status": normalized_status,
            "target_date": target_date,
        }

    return phase_map


def _extract_phase_from_wbs_id(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    return text.split(".")[0]


def _is_done_token(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"âœ…", "done", "complete", "completed", "yes", "y", "true", "1"}


def _is_not_applicable_token(value: object) -> bool:
    text = str(value or "").strip().lower()
    return text in {"n/a", "na", "not applicable", "-"}


def _row_categories_allow_done(ws_wbs, row: int, category_cols: dict[str, int]) -> bool:
    applicable_count = 0
    completed_count = 0

    for category_name, col_idx in category_cols.items():
        value = ws_wbs.cell(row=row, column=col_idx).value
        if _is_not_applicable_token(value):
            continue

        applicable_count += 1
        if _is_done_token(value):
            completed_count += 1

    return applicable_count > 0 and applicable_count == completed_count


def _apply_row_level_statuses(path: Path) -> None:
    wb = load_workbook(path)
    ws_wbs = wb["WBS"]
    ws_milestones = wb["Milestones"]

    phase_map = _phase_status_from_milestones(ws_milestones)

    headers = _header_map(ws_wbs)
    col_wbs_id = headers.get("WBS ID")
    if not col_wbs_id:
        raise RuntimeError(f"WBS sheet missing 'WBS ID' column in {path}")

    col_exec_status = _ensure_column(ws_wbs, headers, "Execution Status")
    col_status_updated = _ensure_column(ws_wbs, headers, "Status Updated On")
    col_started = _ensure_column(ws_wbs, headers, "Started On")
    col_completed = _ensure_column(ws_wbs, headers, "Completed On")
    category_cols = {name: _ensure_column(ws_wbs, headers, name) for name in CATEGORY_COLUMNS}

    today = date.today()

    for r in range(2, ws_wbs.max_row + 1):
        phase = _extract_phase_from_wbs_id(ws_wbs.cell(row=r, column=col_wbs_id).value)
        if not phase:
            continue

        phase_info = phase_map.get(phase, {"status": STATUS_NOT_STARTED, "target_date": None})
        status = str(phase_info["status"])
        target_date = phase_info.get("target_date")

        if status == STATUS_DONE and not _row_categories_allow_done(ws_wbs, r, category_cols):
            status = STATUS_PARTIAL

        ws_wbs.cell(row=r, column=col_exec_status).value = status
        ws_wbs.cell(row=r, column=col_status_updated).value = today
        ws_wbs.cell(row=r, column=col_status_updated).number_format = "yyyy-mm-dd"

        if status == STATUS_DONE:
            ws_wbs.cell(row=r, column=col_started).value = target_date or today
            ws_wbs.cell(row=r, column=col_completed).value = target_date or today
            ws_wbs.cell(row=r, column=col_started).number_format = "yyyy-mm-dd"
            ws_wbs.cell(row=r, column=col_completed).number_format = "yyyy-mm-dd"
        elif status == STATUS_PARTIAL:
            ws_wbs.cell(row=r, column=col_started).value = target_date or today
            ws_wbs.cell(row=r, column=col_started).number_format = "yyyy-mm-dd"
            ws_wbs.cell(row=r, column=col_completed).value = None
        else:
            ws_wbs.cell(row=r, column=col_started).value = None
            ws_wbs.cell(row=r, column=col_completed).value = None

    wb.save(path)
    print(f"updated row-level WBS status: {path}")


if __name__ == "__main__":
    for workbook in FILES:
        _apply_row_level_statuses(workbook)
