from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
FILES = [
    ROOT / "docs" / "YAM_AGRI_WBS_GANTT.xlsx",
    ROOT / "docs" / "planning" / "YAM_AGRI_WBS_GANTT.xlsx",
]

CATEGORY_COLUMNS = ["Schema", "Validation", "Permissions/Isolation", "Workflow", "Evidence"]
DONE_TOKEN = "✅"


def _header_map(ws) -> dict[str, int]:
    return {
        str(cell.value).strip(): idx + 1
        for idx, cell in enumerate(ws[1])
        if cell.value is not None and str(cell.value).strip()
    }


def _ensure_column(ws, headers: dict[str, int], header_name: str) -> int:
    if header_name in headers:
        return headers[header_name]

    col = ws.max_column + 1
    ws.cell(row=1, column=col).value = header_name
    headers[header_name] = col
    return col


def _extract_phase_from_wbs_id(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    return text.split(".")[0]


def _milestone_done_phases(ws_milestones) -> set[str]:
    headers = _header_map(ws_milestones)
    col_milestone = headers.get("Milestone")
    col_status = headers.get("Status")
    if not col_milestone or not col_status:
        raise RuntimeError("Milestones sheet missing Milestone/Status columns")

    done_phases: set[str] = set()
    for r in range(2, ws_milestones.max_row + 1):
        milestone = ws_milestones.cell(row=r, column=col_milestone).value
        status = ws_milestones.cell(row=r, column=col_status).value
        if not milestone:
            continue

        m_text = str(milestone).strip().upper()
        s_text = str(status or "").strip().lower()
        if m_text.startswith("M") and ("✅" in s_text or "done" in s_text):
            done_phases.add(m_text[1:])

    return done_phases


def _is_blank(value: object) -> bool:
    return value is None or not str(value).strip()


def _prefill(path: Path) -> None:
    wb = load_workbook(path)
    ws_wbs = wb["WBS"]
    ws_milestones = wb["Milestones"]

    done_phases = _milestone_done_phases(ws_milestones)

    headers = _header_map(ws_wbs)
    col_wbs_id = headers.get("WBS ID")
    if not col_wbs_id:
        raise RuntimeError(f"WBS sheet missing 'WBS ID' column in {path}")

    category_cols = {name: _ensure_column(ws_wbs, headers, name) for name in CATEGORY_COLUMNS}

    filled = 0
    for r in range(2, ws_wbs.max_row + 1):
        phase = _extract_phase_from_wbs_id(ws_wbs.cell(row=r, column=col_wbs_id).value)
        if not phase or phase not in done_phases:
            continue

        for col in category_cols.values():
            cell = ws_wbs.cell(row=r, column=col)
            if _is_blank(cell.value):
                cell.value = DONE_TOKEN
                filled += 1

    wb.save(path)
    print(f"prefilled categories: {path} (filled={filled})")


if __name__ == "__main__":
    for workbook in FILES:
        _prefill(workbook)
